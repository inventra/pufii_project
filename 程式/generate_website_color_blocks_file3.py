#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""從商品資料(大檔).xlsx 的 商品資料(大檔)+大師 產生 {日期}新品-檔案3-色塊.xlsx。"""
from __future__ import annotations

import argparse, datetime as dt, hashlib, json
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path('/Users/openclaw/創旭/帕妃')
HEADERS = ['物品編號', '顏色', '色塊路徑', '瀏覽用小圖', '大圖路徑', '超大圖路徑', '排序', '銷售']


def norm(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def to_number(v: Any) -> int | None:
    s = norm(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def today_mmdd() -> str:
    return dt.datetime.now().strftime('%m%d')


def current_year() -> str:
    return dt.datetime.now().strftime('%Y')


def find_big_file(input_dir: Path) -> Path:
    candidates = [
        input_dir / '商品資料(大檔).xlsx',
        input_dir / '商品資料大檔.xlsx',
    ] + sorted(input_dir.glob('*商品資料*大檔*.xlsx'))
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f'找不到商品資料大檔，請放在 {input_dir}')


def load_listing_codes(wb, date_code: str) -> list[str]:
    """商品資料(大檔): A 記號=date_code, B 用途=上架, E 流水號。"""
    ws = wb['商品資料(大檔)']
    codes: list[str] = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)
        purpose = norm(ws.cell(r, 2).value)
        code = norm(ws.cell(r, 5).value)
        if mark == date_code and purpose == '上架' and code and code not in seen:
            codes.append(code)
            seen.add(code)
    return codes


def load_master_rows(wb, codes: list[str], date_code: str, year: str) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    """大師: A 物品編號, J 商店顏色, AC 圖片路徑, AD 色塊號碼。

    色塊/中圖以「同一流水號 + 商店顏色」去重，不因尺寸重複產生。
    """
    ws = wb['大師']
    code_set = set(codes)
    order = {c: i for i, c in enumerate(codes)}
    rows_by_key: dict[tuple[str, str], tuple[int, list[Any]]] = {}
    missing: list[dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)
        if code not in code_set:
            continue
        color = norm(ws.cell(r, 10).value)
        image_name = norm(ws.cell(r, 29).value)
        color_no = to_number(ws.cell(r, 30).value)
        if not color:
            missing.append({'row': r, '物品編號': code, 'missing': '商店顏色'})
            continue
        key = (code, color)
        if key in rows_by_key:
            continue
        if not image_name or color_no is None:
            missing.append({
                'row': r,
                '物品編號': code,
                '顏色': color,
                '圖片路徑': image_name,
                '色塊號碼': norm(ws.cell(r, 30).value),
                'missing': '圖片路徑或色塊號碼',
            })
            continue
        color_path = f'https://photo.pufii.com.tw/{year}/{date_code}-C/{image_name}-{color_no}.jpg'
        sort_no = color_no - 2
        rows_by_key[key] = (r, [code, color, color_path, '', '', '', sort_no, 1])

    # 保留大師分頁出現順序；同時依當日流水號順序穩定排序。
    ordered = sorted(rows_by_key.items(), key=lambda item: (order[item[0][0]], item[1][0]))
    rows = [payload for _key, (_r, payload) in ordered]
    return rows, missing


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = len(norm(ws.cell(1, c).value))
        for r in range(2, min(ws.max_row, 100) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 60)


def generate(base: Path, date_code: str, source: Path | None = None, year: str | None = None, dry_run: bool = False) -> Path:
    input_dir = base / '輸入檔'
    output_dir = base / '輸出檔' / f'{date_code}官網上架匯入收音機的檔案'
    output_dir.mkdir(parents=True, exist_ok=True)
    source_path = source or find_big_file(input_dir)
    output_path = output_dir / f'{date_code}新品-檔案3-色塊.xlsx'
    year = year or current_year()

    wb_src = load_workbook(source_path, data_only=True)
    listing_codes = load_listing_codes(wb_src, date_code)
    rows, missing = load_master_rows(wb_src, listing_codes, date_code, year)

    wb = Workbook()
    ws = wb.active
    ws.title = '商品資料'
    ws.append(HEADERS)
    fill = PatternFill('solid', fgColor='D9EAF7')
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    autosize(ws)

    audit = {
        'generated_at': dt.datetime.now().isoformat(timespec='seconds'),
        'script': str(Path(__file__).resolve()),
        'base': str(base),
        'date_code': date_code,
        'year': year,
        'sheet': '商品資料',
        'columns': HEADERS,
        'filter': '商品資料(大檔)!A 記號 = date_code 且 B 用途 = 上架；取 E 流水號',
        'source_sheet': '大師',
        'source_mapping': {
            '物品編號': '大師!A 物品編號',
            '顏色': '大師!J 商店顏色（同流水號+商店顏色去重）',
            '色塊路徑': 'https://photo.pufii.com.tw/{year}/{date}-C/{大師!AC 圖片路徑}-{大師!AD 色塊號碼}.jpg',
            '排序': '大師!AD 色塊號碼 - 2',
            '銷售': '固定 1',
        },
        'listing_code_count': len(listing_codes),
        'listing_codes': listing_codes,
        'row_count': len(rows),
        'missing_count': len(missing),
        'missing': missing[:200],
        'inputs': {'商品資料大檔': {'path': str(source_path), 'sha256': sha256(source_path)}},
        'output': str(output_path),
        'dry_run': dry_run,
    }
    if dry_run:
        print(json.dumps(audit, ensure_ascii=False, indent=2))
        return output_path
    wb.save(output_path)
    audit['output_sha256'] = sha256(output_path)
    Path(str(output_path) + '.audit.json').write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    return output_path


def main() -> None:
    ap = argparse.ArgumentParser(description='產生 {日期}新品-檔案3-色塊')
    ap.add_argument('--base', default=str(DEFAULT_BASE))
    ap.add_argument('--date', default=today_mmdd())
    ap.add_argument('--source', default='')
    ap.add_argument('--year', default='')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    out = generate(Path(args.base), args.date, Path(args.source) if args.source else None, args.year or None, args.dry_run)
    if not args.dry_run:
        print(f'已產出：{out}')
        print(f'產出紀錄：{out}.audit.json')


if __name__ == '__main__':
    main()
