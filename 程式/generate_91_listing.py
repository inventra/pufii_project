#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
帕妃 91-{日期}新品上架的檔案 產生程式（草稿 / 可審核版）

資料來源：
- {base}/輸入檔/商品資料大檔.xlsx
- {base}/輸入檔/尺寸表&試穿報告.xlsx

輸出：
- {base}/輸出檔/91-{日期}新品上架的檔案.xlsx
- {base}/輸出檔/91-{日期}新品上架的檔案.xlsx.audit.json

目前已知規則：
- 分頁名稱：商品資料
- 篩選：商品資料(大檔) B 欄「用途」= 上架
- 商品選項一依 `顏色尺寸請用這一個` 展開：每個有完整顏色+尺寸的選項一列，例如 `顏色:黑F`、`顏色:黑S`；缺資料的選項不輸出空白列，只寫入 audit/待補資料。
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path("/Users/openclaw/創旭/帕妃")

FEATURE_TEXT = """※商品追加需7-30天(不含六日、國定假日)
※訂購前請詳閱賣場【活動說明】和【商店介紹→購物說明】，訂購即表示同意賣場規定
※訂單金額超過4千元，請勿選擇超商取貨，請選宅配
※下單後無法更改取貨門市!
※APP無寄送離島服務，請勿在此下單，離島請聯繫客服!"""

HEADERS = [
    "商品品類",
    "商店類別",
    "商品名稱",
    "數量",
    "建議售價",
    "售價",
    "成本",
    "一次最高購買量",
    "銷售開始日期",
    "銷售結束日期",
    "交期",
    "預定出貨日期",
    "付款完成後幾天出貨",
    "配送方式",
    "付款方式",
    "商品選項",
    "商品選項一",
    "商品選項二",
    "商品料號",
    "商品選項圖檔",
    "商品規格",
    "商品圖檔一",
    "商品圖檔二",
    "商品圖檔三",
    "商品圖檔四",
    "商品圖檔五",
    "商品圖檔六",
    "商品圖檔七",
    "商品圖檔八",
    "商品圖檔九",
    "商品圖檔十",
    "銷售重點",
    "商品特色",
    "詳細說明",
    "商店名稱",
    "SEOTitle",
    "SEOKeywords",
    "SEODescription",
    "溫層類別",
    "商品材積(長x寬x高)",
    "商品重量",
    "商品是否可退貨",
    "指定到貨日",
    "商品備貨天數",
    "可指定的天數長度",
    "可指定的最早到貨日期",
    "可指定的最晚到貨日期",
    "開放指定當天到貨",
    "安全庫存量",
    "隱賣商品",
]

PENDING_RULES = [
    "銷售重點目前用尺寸表適合尺寸；若尺寸表空白會留空並寫入 audit。",
    "商品選項圖檔/商品圖檔欄位尚未確認，先留空。",
]


def norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r"\s+", " ", str(v).strip())


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def today_mmdd() -> str:
    return dt.datetime.now().strftime("%m%d")


def sales_start_from_mmdd(date_code: str, year: int | None = None) -> str:
    if not re.match(r"^\d{4}$", date_code):
        raise ValueError("date must be MMDD, e.g. 0414")
    y = year or dt.datetime.now().year
    return f"{y:04d}/{int(date_code[:2]):02d}/{int(date_code[2:]):02d}"


def build_fit_map(size_path: Path) -> dict[str, str]:
    """尺寸表各分頁：A 欄流水號 -> 適合尺寸。"""
    wb = load_workbook(size_path, data_only=True)
    fit_map: dict[str, str] = {}
    for ws in wb.worksheets:
        headers = [norm(ws.cell(1, c).value).replace(" ", "").replace("\n", "") for c in range(1, ws.max_column + 1)]
        fit_cols = [i + 1 for i, h in enumerate(headers) if h == "適合尺寸"]
        if not fit_cols:
            continue
        fit_col = fit_cols[0]
        current_code = ""
        for r in range(2, ws.max_row + 1):
            raw_code = norm(ws.cell(r, 1).value)
            if raw_code:
                current_code = raw_code
            code = current_code
            if not code or code in fit_map:
                continue
            fit = norm(ws.cell(r, fit_col).value)
            if fit:
                fit_map[code] = fit
    return fit_map


def build_option_map(big_path: Path) -> dict[str, list[str]]:
    """顏色尺寸請用這一個：A 物品編號 -> 91 商品選項一列表。"""
    wb = load_workbook(big_path, data_only=True)
    if "顏色尺寸請用這一個" not in wb.sheetnames:
        return {}
    ws = wb["顏色尺寸請用這一個"]
    option_map: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)  # A 物品編號
        if not code:
            continue
        color = norm(ws.cell(r, 10).value) or norm(ws.cell(r, 8).value)  # J 商店顏色, fallback H 顏色
        size = norm(ws.cell(r, 11).value)  # K 商店尺寸
        # 91 的商品選項一會影響展開行數；缺顏色或缺尺寸時不輸出半套/空白列。
        if not color or not size:
            continue
        option = f"顏色:{color}{size}"
        if code not in option_map:
            option_map[code] = []
            seen[code] = set()
        if option not in seen[code]:
            option_map[code].append(option)
            seen[code].add(option)
    return option_map


def format_jpg_path(value: Any) -> str:
    path = norm(value)
    if not path:
        return ""
    return path if path.lower().endswith(".jpg") else f"{path}.jpg"


def build_image_path_map(big_path: Path) -> dict[str, str]:
    """大師：A 物品編號 -> 91 商品圖檔一。

    使用者確認：商品圖檔一 = 商品資料大檔.xlsx → 大師 → 圖片路徑 + `.jpg`。
    同一流水號多列時取第一個非空圖片路徑。
    """
    wb = load_workbook(big_path, data_only=True)
    if "大師" not in wb.sheetnames:
        return {}
    ws = wb["大師"]
    image_map: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)  # A 物品編號
        if not code or code in image_map:
            continue
        image_path = format_jpg_path(ws.cell(r, 29).value)  # AC 圖片路徑
        if image_path:
            image_map[code] = image_path
    return image_map


def load_listing_rows(big_path: Path, fit_map: dict[str, str], option_map: dict[str, list[str]], image_map: dict[str, str], date_code: str, year: int | None = None) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    wb = load_workbook(big_path, data_only=True)
    ws = wb["商品資料(大檔)"]
    sales_start = sales_start_from_mmdd(date_code, year)
    rows: list[list[Any]] = []
    audit_missing: list[dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)  # A 記號
        usage = norm(ws.cell(r, 2).value)  # B 用途
        if mark != date_code or usage != "上架":
            continue

        code = norm(ws.cell(r, 5).value)  # E 流水號（刪除直播日期欄後）
        full_room_name = norm(ws.cell(r, 26).value)  # Z 完整房名（刪除直播日期欄後）
        origin_price = ws.cell(r, 19).value  # S 原價（刪除直播日期欄後）
        new_price = ws.cell(r, 20).value  # T 新品價（刪除直播日期欄後）
        ray_cost = ws.cell(r, 11).value  # K ray 成本（刪除直播日期欄後）
        fit = fit_map.get(code, "")
        image_path = image_map.get(code, "")
        code_missing = False
        if not full_room_name:
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "商品名稱",
                "reason": "商品資料(大檔) Z欄完整房名空白；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 商品資料(大檔)：對應流水號且 B欄用途=上架 的那一列，Z欄完整房名需有值",
            })
            code_missing = True
        if origin_price is None or norm(origin_price) == "":
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "建議售價",
                "reason": "商品資料(大檔) S欄原價空白；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 商品資料(大檔)：對應流水號且 B欄用途=上架 的那一列，S欄原價需有值",
            })
            code_missing = True
        if new_price is None or norm(new_price) == "":
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "售價",
                "reason": "商品資料(大檔) T欄新品價空白；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 商品資料(大檔)：對應流水號且 B欄用途=上架 的那一列，T欄新品價需有值",
            })
            code_missing = True
        if ray_cost is None or norm(ray_cost) == "":
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "成本",
                "reason": "商品資料(大檔) K欄 ray成本空白；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 商品資料(大檔)：對應流水號且 B欄用途=上架 的那一列，K欄 ray成本需有值",
            })
            code_missing = True
        if not fit:
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "銷售重點/適合尺寸",
                "reason": "尺寸表找不到適合尺寸或適合尺寸空白；91 此流水號先不顯示",
                "source": "尺寸表&試穿報告.xlsx：對應流水號的尺寸/適合尺寸欄需有值",
            })
            code_missing = True
        if not image_path:
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "商品圖檔一",
                "reason": "大師工作表找不到圖片路徑；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 大師：A欄物品編號需有此流水號，AC欄圖片路徑需有值",
            })
            code_missing = True

        # 商品料號：取商品名稱最後一組【】裡面的文字，例如 ...【CP33034】 -> CP33034
        product_no = ""
        bracket_matches = re.findall(r"【([^】]+)】", full_room_name)
        if bracket_matches:
            product_no = bracket_matches[-1].strip()

        options = option_map.get(code, [])
        if not options:
            audit_missing.append({
                "row": r,
                "code": code,
                "field": "商品選項一",
                "reason": "顏色尺寸請用這一個找不到此流水號的完整商店顏色+商店尺寸；91 此流水號先不顯示",
                "source": "商品資料大檔.xlsx → 顏色尺寸請用這一個：A欄物品編號需有此流水號，J欄商店顏色與 K欄商店尺寸需有值",
            })
            code_missing = True

        # 91 上架檔：只要這個流水號有任一必填缺資料，就整個先不顯示。
        # 待補資料.txt 會列出此流水號缺哪些欄位，補完來源後再重跑。
        if code_missing:
            continue

        for option in options:
            row = [
                "服裝、內睡衣 >女裝 >上衣 >T恤/帽T",  # A 商品品類
                "☰ 上衣>T恤、造型上衣",  # B 商店類別
                full_room_name,  # C 商品名稱
                999,  # D 數量
                origin_price,  # E 建議售價
                new_price,  # F 售價
                ray_cost,  # G 成本
                50,  # H 一次最高購買量
                sales_start,  # I 銷售開始日期
                "8/24/41 0:00",  # J 銷售結束日期
                "一般",  # K 交期
                "",  # L 預定出貨日期
                "",  # M 付款完成後幾天出貨
                "6592;101268;101269;99006;99007",  # N 配送方式
                "信用卡一次付款,全家取貨付款,7-11取貨付款,ATM付款,LINE Pay,街口支付,Apple Pay,AFTEE先享後付,悠遊付",  # O 付款方式
                "有",  # P 商品選項
                option, "", product_no, "", "",  # Q-U 商品選項一/二/料號/選項圖檔/規格
                image_path, "", "", "", "", "", "", "", "", "",  # V-AE 商品圖檔一～十
                fit,  # AF 銷售重點
                FEATURE_TEXT,  # AG 商品特色
                "",  # AH 詳細說明
                "PUFII-APP",  # AI 商店名稱
                "",  # AJ SEOTitle
                "",  # AK SEOKeywords
                "",  # AL SEODescription
                "常溫",  # AM 溫層類別
                "",  # AN 商品材積
                "",  # AO 商品重量
                "可退貨",  # AP 商品是否可退貨
                "關閉",  # AQ 指定到貨日
                "",  # AR 商品備貨天數
                "",  # AS 可指定的天數長度
                "",  # AT 可指定的最早到貨日期
                "",  # AU 可指定的最晚到貨日期
                "",  # AV 開放指定當天到貨
                0,  # AW 安全庫存量
                "是",  # AX 隱賣商品
            ]
            if len(row) != len(HEADERS):
                raise RuntimeError(f"row length {len(row)} != headers {len(HEADERS)}")
            rows.append(row)
    return rows, audit_missing


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = len(norm(ws.cell(1, c).value))
        for r in range(2, min(ws.max_row, 80) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value).split("\n")[0]))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 8), 35)


def generate(base: Path, date_code: str, year: int | None = None, dry_run: bool = False) -> Path:
    input_dir = base / "輸入檔"
    output_dir = base / "輸出檔"
    output_dir.mkdir(parents=True, exist_ok=True)

    big_path = input_dir / "商品資料大檔.xlsx"
    size_path = input_dir / "尺寸表&試穿報告.xlsx"
    output_path = output_dir / f"91-{date_code}新品上架的檔案.xlsx"

    fit_map = build_fit_map(size_path)
    option_map = build_option_map(big_path)
    image_map = build_image_path_map(big_path)
    rows, audit_missing = load_listing_rows(big_path, fit_map, option_map, image_map, date_code, year)

    wb = Workbook()
    ws = wb.active
    ws.title = "商品資料"
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)
    ws.column_dimensions["AG"].width = 60
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 33).alignment = Alignment(wrap_text=True, vertical="top")

    audit = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "script": str(Path(__file__).resolve()),
        "base": str(base),
        "date_code": date_code,
        "year": year or dt.datetime.now().year,
        "sheet": "商品資料",
        "filter": "商品資料(大檔)!A 記號 = date_code 且 B 用途 = 上架",
        "column_count": len(HEADERS),
        "last_column": get_column_letter(len(HEADERS)),
        "last_header": HEADERS[-1],
        "row_count": len(rows),
        "inputs": {
            "商品資料大檔.xlsx": {"path": str(big_path), "sha256": sha256(big_path)},
            "尺寸表&試穿報告.xlsx": {"path": str(size_path), "sha256": sha256(size_path)},
        },
        "output": str(output_path),
        "pending_rules": PENDING_RULES,
        "missing_or_pending": audit_missing,
        "dry_run": dry_run,
    }

    if dry_run:
        print(json.dumps(audit, ensure_ascii=False, indent=2))
        return output_path

    wb.save(output_path)
    audit["output_sha256"] = sha256(output_path)
    Path(str(output_path) + ".audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="產生 91-{日期}新品上架的檔案.xlsx")
    parser.add_argument("--base", default=str(DEFAULT_BASE), help="帕妃資料夾根目錄")
    parser.add_argument("--date", default=today_mmdd(), help="日期 MMDD，例如 0414；未填預設今天")
    parser.add_argument("--year", type=int, default=dt.datetime.now().year, help="銷售開始日期年份，預設今年")
    parser.add_argument("--dry-run", action="store_true", help="只輸出 audit，不寫入 Excel")
    args = parser.parse_args()
    output = generate(Path(args.base), args.date, args.year, args.dry_run)
    if not args.dry_run:
        print(f"已產出：{output}")
        print(f"產出紀錄：{output}.audit.json")


if __name__ == "__main__":
    main()
