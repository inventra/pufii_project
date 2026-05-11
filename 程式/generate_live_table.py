#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
帕妃直播表格產生程式（草稿 / 可審核版）

用途：
- 從 商品資料大檔.xlsx + 尺寸表&試穿報告.xlsx 產出 {日期}直播表格.xlsx
- 目前先整理已確認規則，尚未確認的欄位先留空或列入 audit

如何證明檔案是程式產出的：
- 每次執行會同時產出 audit JSON：輸出檔名 + .audit.json
- audit 會記錄：執行時間、script 路徑、輸入檔 SHA256、輸出檔 SHA256、使用日期、待確認項目
- 可用 `python3 generate_live_table.py --date 0414` 重新跑一次，比對 SHA256 或內容

注意：
- 使用者規則：檔名不要亂改，同一份輸出用迭代覆蓋，除非使用者指定版本區分。
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from excel_highlight_utils import highlight_cell_if_blank, header_map


DEFAULT_BASE = Path("/Users/openclaw/創旭/帕妃")


MAIN_HEADERS = [
    "備註", "貨號", "標數", "{date}直播", "推薦尺寸", "顏色",
    "原價", "直播價", "尺寸表", "成本", "新品價", "廠商",
]
MATCH_HEADERS = ["備註", "貨號", "標數", "{date}直播", "推薦尺寸", "顏色"]
SISTER_HEADERS = ["標數", "{date}直播", "推薦尺寸", "顏色"]
RAY_HEADERS = [
    "備註", "貨號", "標數", "{date}直播", "推薦尺寸", "顏色",
    "原價", "直播價", "成本", "新品價", "廠商",
]


# 待客戶確認：這些欄位暫時不自動定案
PENDING_RULES = [
    "推薦尺寸若尺寸表空白但正確檔有值，來源仍需確認。",
    "P33315 品名差異需確認：主表/RAY 與對款名稱不同，且有『背心/肩帶背心』差異。",
]


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def header_key(value: Any) -> str:
    """比對欄名用：去掉換行與空白。"""
    return norm(value).replace(" ", "").replace("\n", "")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def mmdd_today() -> str:
    return dt.datetime.now().strftime("%m%d")


def find_header(headers: list[str], candidates: list[str], allow_prefix: bool = False) -> int | None:
    """回傳 1-based column index。

    allow_prefix 用於尺寸欄：例如來源表頭「腰寬\n鬆緊帶x~x」要可用「腰寬」找到。
    """
    keys = [h.replace(" ", "").replace("\n", "") for h in headers]
    for candidate in candidates:
        ck = candidate.replace(" ", "").replace("\n", "")
        for i, h in enumerate(keys, start=1):
            if h == ck or (allow_prefix and h.startswith(ck)):
                return i
    return None


def lookup_big_rows(big_path: Path, date_code: str) -> list[dict[str, Any]]:
    """從 商品資料(大檔) 篩出記號 = date_code 的商品（新版欄位：A記號、B用途、E流水號、S/T價格）。"""
    wb = load_workbook(big_path, data_only=True)
    ws = wb["商品資料(大檔)"]
    rows = []
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)  # A 記號
        purpose = norm(ws.cell(r, 2).value)  # B 用途（新版已刪直播日期欄）
        if mark != date_code or purpose != "上架":
            continue
        rows.append({
            "row": r,
            "code": norm(ws.cell(r, 5).value),       # E 流水號
            "sku": norm(ws.cell(r, 7).value),        # G 貨號
            "domestic_vendor": norm(ws.cell(r, 8).value),
            "domestic_cost": ws.cell(r, 9).value,
            "vendor_sku": norm(ws.cell(r, 10).value),
            "ray_cost": ws.cell(r, 11).value,         # K ray成本
            "foreign_vendor": norm(ws.cell(r, 12).value),
            "foreign_place": norm(ws.cell(r, 13).value),
            "name_short": norm(ws.cell(r, 15).value),
            "room_name": norm(ws.cell(r, 17).value),  # Q 房名
            "origin_price_big": ws.cell(r, 19).value, # S 原價
            "new_price_big": ws.cell(r, 20).value,    # T 新品價
            "live_price_big": ws.cell(r, 22).value,   # V 直播+1價
            "purpose": purpose,
        })
    return rows


def lookup_ray_prices(big_path: Path) -> dict[str, dict[str, Any]]:
    """ray價錢：用 I 欄流水號對應，取 K/M/L/O 等欄。"""
    wb = load_workbook(big_path, data_only=True)
    ws = wb["ray價錢"]
    data = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 9).value)  # I 流水號
        if not code:
            continue
        data[code] = {
            "row": r,
            "cost_k": ws.cell(r, 11).value,       # K 成本(台幣)
            "origin_l": ws.cell(r, 12).value,     # L 原價
            "new_m": ws.cell(r, 13).value,        # M 新品價
            "live_o": ws.cell(r, 15).value,       # O 直播+1價
            "foreign_place_q": norm(ws.cell(r, 17).value),
            "foreign_cost_r": ws.cell(r, 18).value,
            "weight_s": ws.cell(r, 19).value,
        }
    return data


def lookup_color_map(big_path: Path) -> dict[str, str]:
    """顏色來源：商品資料大檔 → 顏色尺寸請用這一個，依物品編號取 J 商店顏色去重合併。"""
    wb = load_workbook(big_path, data_only=True)
    if "顏色尺寸請用這一個" not in wb.sheetnames:
        return {}
    ws = wb["顏色尺寸請用這一個"]
    data: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)  # A 物品編號
        color = norm(ws.cell(r, 10).value) or norm(ws.cell(r, 8).value)  # J 商店顏色，備援 H 顏色
        if not code or not color:
            continue
        seen.setdefault(code, set())
        data.setdefault(code, [])
        if color not in seen[code]:
            seen[code].add(color)
            data[code].append(color)
    return {code: "、".join(colors) for code, colors in data.items()}


SIZE_ORDER = ["XS", "S", "M", "L", "XL", "2L", "3L", "4L", "5L"]
SIZE_RANK = {size: i for i, size in enumerate(SIZE_ORDER)}


def extract_variant_size(value: str) -> str:
    """從 C欄顏色/尺寸中抓實際拿貨尺寸，例如 藍/白S -> S。"""
    text = norm(value).upper().replace(" ", "")
    matches = re.findall(r"(?:[2-5]L|XL|XS|S|M|L)", text)
    return matches[-1] if matches else norm(value)


def normalize_body_size_text(value: str) -> str:
    text = norm(value).replace("適合", "").upper()
    text = text.replace("／", "/").replace("、", "/").replace(".", "/")
    parts = [p for p in re.split(r"[/]+", text) if p]
    return ".".join(parts)


def size_tokens(value: str) -> list[str]:
    return re.findall(r"(?:[2-5]L|XL|XS|S|M|L)", value.upper())


def format_recommended_fit(fit_rows: list[dict[str, str]], fallback_values: list[str]) -> str:
    """多尺寸推薦尺寸格式。

    例：C欄 S + D欄 S-S/M適合、C欄 L + D欄 L-L/XL適合
    -> S-XL適合\nS.M拿S\nL.XL拿L
    """
    parsed = []
    all_body_tokens = []
    for row in fit_rows:
        variant = extract_variant_size(row.get("variant", ""))
        fit = norm(row.get("fit", ""))
        if not variant or not fit:
            continue
        fit_body = fit.replace("適合", "")
        if "-" in fit_body:
            fit_body = fit_body.split("-", 1)[1]
        elif "－" in fit_body:
            fit_body = fit_body.split("－", 1)[1]
        body_display = normalize_body_size_text(fit_body)
        tokens = size_tokens(body_display)
        if not body_display or not tokens:
            continue
        all_body_tokens.extend(tokens)
        parsed.append((body_display, variant))

    if len(parsed) <= 1 or not all_body_tokens:
        return " / ".join(fallback_values)

    ranked = [t for t in all_body_tokens if t in SIZE_RANK]
    if ranked:
        start = min(ranked, key=lambda t: SIZE_RANK[t])
        end = max(ranked, key=lambda t: SIZE_RANK[t])
        summary = f"{start}-{end}適合"
    else:
        summary = " / ".join(fallback_values)
    details = [f"{body}拿{variant}" for body, variant in parsed]
    return "\n".join([summary] + details)


def build_size_maps(size_path: Path) -> dict[str, dict[str, Any]]:
    """建立 流水號 -> 尺寸文字/推薦尺寸 的 map。

    重要規則：
    - 欄位名稱原則上保留來源名稱，不自行把 大腿 改成 大腿寬、下擺 改成 下擺寬。
    - 上衣+下身分頁若找到，第二段下身需額外讀 Q-W 欄。
    - 同一商品不同尺寸可只有第一列填流水號；後續 A欄空白列沿用上一個流水號，直到下一個流水號出現。
    - 空白欄位不輸出。
    """
    wb = load_workbook(size_path, data_only=True)
    result: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        headers_raw = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        headers_key = [header_key(h) for h in headers_raw]
        fit_col = find_header(headers_raw, ["適合尺寸", "適合\n尺寸"])
        current_code = ""

        for r in range(2, ws.max_row + 1):
            raw_code = norm(ws.cell(r, 1).value)
            if raw_code:
                if re.match(r"^P\d{5}$", raw_code):
                    current_code = raw_code
                else:
                    current_code = ""
                    continue
            if not current_code:
                continue

            row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            size_text = build_size_text_for_row(ws.title, headers_raw, headers_key, row_values)
            fit = norm(ws.cell(r, fit_col).value) if fit_col else ""
            fit = f"{fit}適合" if fit and not fit.endswith("適合") else fit
            variant = norm(ws.cell(r, 3).value)  # 顏色/尺寸欄，常見為 S/M/L/XL

            entry = result.setdefault(current_code, {
                "sheet": ws.title,
                "rows": [],
                "fit_rows": [],
                "fit_values": [],
                "size_rows": [],
                "fit": "",
                "size_text": "",
            })
            entry["rows"].append(r)
            if fit and fit not in entry["fit_values"]:
                entry["fit_values"].append(fit)
            if fit:
                entry["fit_rows"].append({"variant": variant, "fit": fit})
            if size_text:
                entry["size_rows"].append({"variant": variant, "text": size_text})

    for code, entry in result.items():
        rows = entry.get("size_rows", [])
        if len(rows) > 1:
            # 多尺寸列前綴使用空格，不使用冒號：例如「S 腰寬...」而不是「S：腰寬...」。
            entry["size_text"] = "\n".join(
                f"{row['variant']} {row['text']}" if row.get("variant") else row["text"]
                for row in rows
            )
        elif rows:
            entry["size_text"] = rows[0]["text"]
        entry["fit"] = format_recommended_fit(entry.get("fit_rows", []), entry.get("fit_values", []))
        entry["row"] = entry.get("rows", [None])[0]
    return result


def value_at(row_values: list[Any], col_1based: int) -> str:
    if col_1based <= 0 or col_1based > len(row_values):
        return ""
    return norm(row_values[col_1based - 1])


def append_field(parts: list[str], label: str, value: str) -> None:
    value = norm(value)
    if not value:
        return
    # 尺寸數值中的半形/全形括號拿掉：例如（17-29）或 (17-29) → 17-29
    value = value.replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    clean_label = norm(label)
    if not clean_label:
        return

    # 特殊清理：肩帶欄位像「肩帶(17-29)」要輸出「肩帶17-29」
    if "肩帶" in clean_label:
        clean_label = "肩帶"
        value = value.replace("(", "").replace(")", "")

    # 已確認：袖長(領-袖) 輸出成 袖長
    if "袖長" in clean_label:
        clean_label = "袖長"

    # 已確認：領口 寬/領口寬 統一輸出領口寬
    if clean_label.replace(" ", "") in ["領口寬", "領口"]:
        clean_label = "領口寬"

    # 腰寬欄有時表頭是「腰寬\n鬆緊帶x~x」，輸出只保留腰寬
    if clean_label.replace(" ", "").startswith("腰寬"):
        clean_label = "腰寬"

    # 其他欄位保留來源名稱，不自行把 大腿 改大腿寬、下擺改下擺寬
    clean_label = clean_label.replace(" ", "")
    parts.append(f"{clean_label}{value}")


def build_size_text_for_row(sheet_name: str, headers_raw: list[str], headers_key: list[str], row_values: list[Any]) -> str:
    first_parts: list[str] = []

    # 常見第一段欄位：依欄名找，保留來源命名
    first_candidates = [
        "領口寬", "領口", "肩寬", "袖長(領-袖)", "袖長", "袖寬",
        "胸寬", "腰寬", "臀寬", "大腿", "褲檔", "下擺", "下擺寬", "全長", "褲管",
    ]
    used_cols = set()
    for candidate in first_candidates:
        col = find_header(headers_raw, [candidate], allow_prefix=True)
        if col and col not in used_cols:
            append_field(first_parts, headers_raw[col - 1], value_at(row_values, col))
            used_cols.add(col)

    # 上衣+下身：第二段固定多讀 Q-W（17-23）
    if sheet_name == "上衣+下身":
        second_parts: list[str] = []
        for col in range(17, 24):  # Q-W
            label = headers_raw[col - 1] if col - 1 < len(headers_raw) else ""
            append_field(second_parts, label, value_at(row_values, col))
        return " // ".join([p for p in [" ".join(first_parts), " ".join(second_parts)] if p])

    # 上衣+小可愛(兩件式)：第二段目前依已知欄位 Q-U 附近讀出小可愛尺寸
    if sheet_name == "上衣+小可愛(兩件式)":
        second_parts = []
        for col in range(17, 22):  # Q-U
            label = headers_raw[col - 1] if col - 1 < len(headers_raw) else ""
            append_field(second_parts, label, value_at(row_values, col))
        # 正確檔目前 // 後不加空格；此處先用 // 直接串，之後可依客戶要求調整
        if first_parts and second_parts:
            return " ".join(first_parts) + " //" + " ".join(second_parts)
        return " ".join(first_parts or second_parts)

    return " ".join(first_parts)


def write_headers(ws, headers: list[str], date_code: str) -> None:
    final_headers = [h.format(date=date_code) for h in headers]
    ws.append(final_headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, len(final_headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = 8
        for r in range(1, min(ws.max_row, 80) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 45)


def highlight_missing_cells(wb: Workbook, audit_missing: list[dict[str, Any]]) -> int:
    """把 audit 判定缺資料、且輸出仍為空白的欄位反黃。

    只標需要人工檢查/補資料的欄位；像「備註」「標數」這類已確認固定空白不標。
    """
    if not audit_missing:
        return 0

    fields_by_code: dict[str, set[str]] = {}
    for item in audit_missing:
        code = norm(item.get("code"))
        field = norm(item.get("field"))
        if code and field:
            fields_by_code.setdefault(code, set()).add(field)

    highlight_count = 0
    target_fields = {"推薦尺寸", "顏色", "尺寸表", "成本"}
    sheets_with_code = [
        wb.worksheets[0],  # 日期主表：B欄貨號為流水號
        wb["對款"],
        wb["RAY"],
    ]
    for ws in sheets_with_code:
        headers = header_map(ws)
        for row in range(2, ws.max_row + 1):
            code = norm(ws.cell(row, 2).value)
            missing_fields = fields_by_code.get(code, set())
            for field in target_fields & missing_fields:
                col = headers.get(field)
                if col and highlight_cell_if_blank(ws, row, col):
                    highlight_count += 1

    # 「姐」分頁沒有流水號欄，依主表同列順序對應，只標推薦尺寸/顏色。
    ws_sister = wb["姐"]
    sister_headers = header_map(ws_sister)
    ws_main = wb.worksheets[0]
    for row in range(2, ws_sister.max_row + 1):
        code = norm(ws_main.cell(row, 2).value)
        missing_fields = fields_by_code.get(code, set())
        for field in {"推薦尺寸", "顏色"} & missing_fields:
            col = sister_headers.get(field)
            if col and highlight_cell_if_blank(ws_sister, row, col):
                highlight_count += 1

    return highlight_count


def write_missing_report(report_path: Path, date_code: str, missing_items: list[dict[str, Any]]) -> None:
    """輸出人工可讀 TXT：主軸按流水號列出需要補哪些欄位；不靠顏色辨識流水號。"""
    lines = [
        f"{date_code}直播表格 待補資料",
        "=" * 40,
        "說明：以下只列會造成輸出資料不完整的來源缺漏；廠商欄位對應商品資料(大檔) H欄，允許原始空白，不列為錯誤。",
        "重要：商品辨識一律用流水號 / 物品編號；顏色只是要補的欄位，不用顏色去認列流水號。",
        "",
    ]
    if not missing_items:
        lines.append("目前沒有偵測到會造成輸出不完整的原始漏資料。")
        report_path.write_text("\n".join(lines), encoding="utf-8")
        return

    by_code: dict[str, list[dict[str, Any]]] = {}
    for item in missing_items:
        code = item.get("code", "未標示流水號")
        by_code.setdefault(code, []).append(item)

    lines.append("一、依流水號整理需補欄位")
    lines.append("-" * 40)
    for idx, (code, items) in enumerate(by_code.items(), start=1):
        fields = []
        for item in items:
            field = item.get("field", "")
            if field and field not in fields:
                fields.append(field)
        lines.append(f"{idx}. 流水號：{code}")
        lines.append(f"   需要補：{'、'.join(fields)}")
        for item in items:
            field = item.get("field", "")
            reason = item.get("reason", "")
            source = item.get("source", "")
            lines.append(f"   - {field}：{reason}")
            if source:
                lines.append(f"     補資料位置：{source}")
        lines.append("")

    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in missing_items:
        source = item.get("source", "未標示來源")
        grouped.setdefault(source, []).append(item)

    lines.append("二、依原始來源彙總")
    lines.append("-" * 40)
    for idx, (source, items) in enumerate(grouped.items(), start=1):
        codes = []
        fields = []
        for item in items:
            code = item.get("code", "")
            field = item.get("field", "")
            if code and code not in codes:
                codes.append(code)
            if field and field not in fields:
                fields.append(field)
        lines.append(f"{idx}. 來源：{source}")
        lines.append(f"   影響欄位：{'、'.join(fields)}")
        lines.append(f"   影響流水號：{'、'.join(codes)}")
        lines.append("")
    report_path.write_text("\n".join(lines), encoding="utf-8")


def generate(base: Path, date_code: str, dry_run: bool = False) -> Path:
    input_dir = base / "輸入檔"
    output_dir = base / "輸出檔"
    output_dir.mkdir(parents=True, exist_ok=True)

    big_path = input_dir / "商品資料大檔.xlsx"
    size_path = input_dir / "尺寸表&試穿報告.xlsx"
    output_path = output_dir / f"{date_code}直播表格.xlsx"

    rows = lookup_big_rows(big_path, date_code)
    ray_map = lookup_ray_prices(big_path)
    color_map = lookup_color_map(big_path)
    size_map = build_size_maps(size_path)

    wb = Workbook()
    wb.remove(wb.active)
    ws_main = wb.create_sheet(date_code)
    ws_match = wb.create_sheet("對款")
    ws_sister = wb.create_sheet("姐")
    ws_ray = wb.create_sheet("RAY")

    write_headers(ws_main, MAIN_HEADERS, date_code)
    write_headers(ws_match, MATCH_HEADERS, date_code)
    write_headers(ws_sister, SISTER_HEADERS, date_code)
    write_headers(ws_ray, RAY_HEADERS, date_code)

    audit_missing = []

    for item in rows:
        code = item["code"]
        ray = ray_map.get(code, {})
        size = size_map.get(code, {})

        # 已確認：標數空白
        mark_no = ""

        # 顏色：商品資料大檔 → 顏色尺寸請用這一個 → J 商店顏色去重合併
        color = color_map.get(code, "")
        if not color:
            audit_missing.append({
                "code": code,
                "field": "顏色",
                "reason": "顏色尺寸請用這一個找不到商店顏色",
                "source": "商品資料大檔.xlsx → 顏色尺寸請用這一個：A欄物品編號需有此流水號，J欄商店顏色需有值",
            })

        # 價格：直接使用大檔新版欄位 S 原價、V 直播+1價、T 新品價
        origin_price = item["origin_price_big"]
        live_price = item["live_price_big"]
        new_price = item["new_price_big"]

        # 成本：直接使用大檔 K ray成本
        cost_candidate = item["ray_cost"]
        if cost_candidate is None or norm(cost_candidate) == "":
            audit_missing.append({
                "code": code,
                "field": "成本",
                "reason": "商品資料(大檔) K欄 ray成本空白，導致直播表格成本空白",
                "source": "商品資料大檔.xlsx → 商品資料(大檔)：對應流水號且 B欄用途=上架 的那一列，K欄 ray成本需有值",
            })

        fit = size.get("fit", "")
        size_text = size.get("size_text", "")
        if not size_text:
            audit_missing.append({
                "code": code,
                "field": "尺寸表",
                "reason": "尺寸表找不到或組不出文字",
                "source": "尺寸表&試穿報告.xlsx：各尺寸分頁 A欄流水號需有此商品，且尺寸欄位需有值",
            })
        if not fit:
            audit_missing.append({
                "code": code,
                "field": "推薦尺寸",
                "reason": "尺寸表適合尺寸空白或找不到",
                "source": "尺寸表&試穿報告.xlsx：對應流水號所在列的『適合尺寸』欄需有值",
            })

        # 廠商：使用商品資料(大檔) H 欄（國內廠商代號/藍色欄位）；允許空白，不列為缺資料
        vendor = item["domestic_vendor"]

        room_name = item["room_name"]
        # 待確認：P33315 在不同分頁名稱不同，目前仍用 room_name，不硬改

        ws_main.append([
            "", code, mark_no, room_name, fit, color,
            origin_price, live_price, size_text, cost_candidate, new_price, vendor,
        ])
        ws_match.append(["", code, mark_no, room_name, fit, color])
        ws_sister.append([mark_no, room_name, fit, color])
        ws_ray.append([
            "", code, mark_no, room_name, fit, color,
            origin_price, live_price, cost_candidate, new_price, vendor,
        ])

    highlight_count = highlight_missing_cells(wb, audit_missing)

    for ws in wb.worksheets:
        autosize(ws)

    audit = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "script": str(Path(__file__).resolve()),
        "base": str(base),
        "date_code": date_code,
        "inputs": {
            "商品資料大檔.xlsx": {"path": str(big_path), "sha256": sha256(big_path)},
            "尺寸表&試穿報告.xlsx": {"path": str(size_path), "sha256": sha256(size_path)},
        },
        "output": str(output_path),
        "row_count": len(rows),
        "pending_rules": PENDING_RULES,
        "missing_or_pending": audit_missing,
        "highlighted_blank_cell_count": highlight_count,
        "dry_run": dry_run,
    }

    if not dry_run:
        wb.save(output_path)
        report_path = output_dir / "待補資料.txt"
        write_missing_report(report_path, date_code, audit_missing)
        audit["output_sha256"] = sha256(output_path)
        audit["missing_report"] = str(report_path)
        audit["missing_report_sha256"] = sha256(report_path)
        audit_path = Path(str(output_path) + ".audit.json")
        audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
        return output_path

    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="產生帕妃 {日期}直播表格.xlsx")
    parser.add_argument("--base", default=str(DEFAULT_BASE), help="帕妃資料夾根目錄")
    parser.add_argument("--date", default=mmdd_today(), help="日期 MMDD，例如 0414；未填預設今天")
    parser.add_argument("--dry-run", action="store_true", help="只輸出 audit，不寫入 Excel")
    args = parser.parse_args()

    output = generate(Path(args.base), args.date, args.dry_run)
    if not args.dry_run:
        print(f"已產出：{output}")
        print(f"產出紀錄：{output}.audit.json")


if __name__ == "__main__":
    main()
