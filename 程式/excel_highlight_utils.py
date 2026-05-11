# -*- coding: utf-8 -*-
"""Excel 空白/待補欄位反黃共用工具。"""
from __future__ import annotations

from typing import Any, Iterable

from openpyxl.styles import PatternFill

MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")  # 淡黃：人工需檢查/補資料


def is_blank(value: Any) -> bool:
    """判斷 Excel cell 是否視為空白。"""
    return value is None or str(value).strip() == ""


def header_map(ws, header_row: int = 1) -> dict[str, int]:
    """建立 表頭文字 -> 欄號 的 map。"""
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is not None:
            result[str(value).strip()] = col
    return result


def highlight_blank_cells(ws, columns: Iterable[int | str], start_row: int = 2) -> int:
    """將指定欄位中空白 cell 反黃，回傳反黃數量。

    columns 可傳 1-based 欄號，或直接傳表頭名稱。
    """
    headers = header_map(ws)
    column_numbers: list[int] = []
    for col in columns:
        if isinstance(col, int):
            column_numbers.append(col)
        else:
            found = headers.get(str(col).strip())
            if found:
                column_numbers.append(found)

    count = 0
    for row in range(start_row, ws.max_row + 1):
        for col in column_numbers:
            cell = ws.cell(row, col)
            if is_blank(cell.value):
                cell.fill = MISSING_FILL
                count += 1
    return count


def highlight_cell_if_blank(ws, row: int, col: int) -> bool:
    """若指定 cell 為空白則反黃。"""
    cell = ws.cell(row, col)
    if is_blank(cell.value):
        cell.fill = MISSING_FILL
        return True
    return False
