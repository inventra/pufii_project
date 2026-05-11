#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
帕妃 {日期}新品-推薦(待填寫) 產生程式

輸入：
- {base}/輸入檔/商品資料大檔.xlsx

輸出：
- {base}/輸出檔/{日期}官網上架匯入收音機的檔案/{日期}新品-推薦(待填寫).xlsx
- 同路徑 .audit.json

規則：
- 商品資料(大檔) 篩選 A欄「記號」= 日期、B欄「用途」= 上架
- 主物品編號 = E欄「流水號」（刪除直播日期欄後）
- 推薦物品編號 = 空白
- 排序 = 空白
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path('/Users/openclaw/創旭/帕妃')
HEADERS = ['主物品編號', '推薦物品編號', '排序']


def norm(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r'\s+', ' ', str(v).strip())


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def today_mmdd() -> str:
    return dt.datetime.now().strftime('%m%d')


def load_rows(big_path: Path, date_code: str) -> list[dict[str, Any]]:
    wb = load_workbook(big_path, data_only=True)
    ws = wb['商品資料(大檔)']
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)  # A 記號
        purpose = norm(ws.cell(r, 2).value)  # B 用途
        code = norm(ws.cell(r, 5).value)  # E 流水號（刪除直播日期欄後）
        if mark == date_code and purpose == '上架' and code:
            rows.append({'source_row': r, 'code': code})
    return rows


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = len(norm(ws.cell(1, c).value))
        for r in range(2, min(ws.max_row, 100) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 12), 30)


def generate(base: Path, date_code: str, dry_run: bool = False) -> Path:
    input_dir = base / '輸入檔'
    output_dir = base / '輸出檔' / f'{date_code}官網上架匯入收音機的檔案'
    output_dir.mkdir(parents=True, exist_ok=True)

    big_path = input_dir / '商品資料大檔.xlsx'
    output_path = output_dir / f'{date_code}新品-推薦(待填寫).xlsx'

    rows = load_rows(big_path, date_code)

    wb = Workbook()
    ws = wb.active
    ws.title = '商品資料'
    ws.append(HEADERS)

    fill = PatternFill('solid', fgColor='D9EAF7')
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in rows:
        ws.append([row['code'], '', ''])

    ws.freeze_panes = 'A2'
    autosize(ws)

    audit = {
        'generated_at': dt.datetime.now().isoformat(timespec='seconds'),
        'script': str(Path(__file__).resolve()),
        'base': str(base),
        'date_code': date_code,
        'folder': str(output_dir),
        'sheet': '商品資料',
        'filter': "商品資料(大檔)!A 記號 = date_code 且 B 用途 = 上架",
        'columns': HEADERS,
        'row_count': len(rows),
        'source_rows': rows,
        'inputs': {'商品資料大檔.xlsx': {'path': str(big_path), 'sha256': sha256(big_path)}},
        'output': str(output_path),
        'dry_run': dry_run,
    }

    if dry_run:
        print(json.dumps(audit, ensure_ascii=False, indent=2))
        return output_path

    wb.save(output_path)
    audit['output_sha256'] = sha256(output_path)
    Path(str(output_path) + '.audit.json').write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description='產生 {日期}新品-推薦(待填寫)')
    parser.add_argument('--base', default=str(DEFAULT_BASE), help='帕妃資料夾根目錄')
    parser.add_argument('--date', default=today_mmdd(), help='日期 MMDD，例如 0414；未填預設今天')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    output = generate(Path(args.base), args.date, args.dry_run)
    if not args.dry_run:
        print(f'已產出：{output}')
        print(f'產出紀錄：{output}.audit.json')


if __name__ == '__main__':
    main()
