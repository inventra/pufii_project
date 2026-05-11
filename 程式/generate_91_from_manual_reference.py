#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""用人工 91 檔的商品料號/選項回推商品資料大檔，產生測試版 91 並輸出差異報告。

用途：當大檔 A/B 日期/用途欄沒有標記，但人工 91 已有完整商品清單時，
用人工 91 的 CPxxxxx -> Pxxxxx 反查商品資料(大檔)，驗證大檔欄位映射與程式輸出。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path("/Users/openclaw/創旭/帕妃")

FEATURE_TEXT = """※商品追加需7-30天(不含六日、國定假日)
※訂購前請詳閱賣場【活動說明】和【商店介紹→購物說明】，訂購即表示同意賣場規定
※訂單金額超過4千元，請勿選擇超商取貨，請選宅配
※下單後無法更改取貨門市!
※APP無寄送離島服務，請勿在此下單，離島請聯繫客服!"""

HEADERS = [
    "商品品類", "商店類別", "商品名稱", "數量", "建議售價", "售價", "成本", "一次最高購買量",
    "銷售開始日期", "銷售結束日期", "交期", "預定出貨日期", "付款完成後幾天出貨", "配送方式", "付款方式",
    "商品選項", "商品選項一", "商品選項二", "商品料號", "商品選項圖檔", "商品規格",
    "商品圖檔一", "商品圖檔二", "商品圖檔三", "商品圖檔四", "商品圖檔五", "商品圖檔六", "商品圖檔七",
    "商品圖檔八", "商品圖檔九", "商品圖檔十", "銷售重點", "商品特色", "詳細說明", "商店名稱", "SEOTitle",
    "SEOKeywords", "SEODescription", "溫層類別", "商品材積(長x寬x高)", "商品重量", "商品是否可退貨", "指定到貨日",
    "商品備貨天數", "可指定的天數長度", "可指定的最早到貨日期", "可指定的最晚到貨日期", "開放指定當天到貨",
    "安全庫存量", "隱賣商品",
]

COMPARE_COLUMNS = [
    "商品名稱", "建議售價", "售價", "成本", "銷售開始日期", "銷售結束日期", "商品選項一", "商品料號", "銷售重點",
]


def norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if hasattr(v, "strftime"):
        # 人工檔日期常是 datetime，統一成 yyyy/m/d 便於比對。
        return f"{v.year}/{v.month}/{v.day}"
    return re.sub(r"\s+", " ", str(v).strip())


def sales_start_from_mmdd(date_code: str, year: int) -> str:
    return f"{year:04d}/{int(date_code[:2])}/{int(date_code[2:])}"


def parse_manual_91(path: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["商品資料"] if "商品資料" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows: list[dict[str, Any]] = []
    options: dict[str, list[str]] = defaultdict(list)
    seen_options: dict[str, set[str]] = defaultdict(set)
    for raw in ws.iter_rows(min_row=2, values_only=True):
        vals = {h: raw[i] if i < len(raw) else None for h, i in idx.items()}
        name = norm(vals.get("商品名稱"))
        product_no = norm(vals.get("商品料號"))
        if not product_no:
            m = re.search(r"【(CP\d+)】", name)
            product_no = m.group(1) if m else ""
        if not product_no.startswith("CP"):
            continue
        code = "P" + product_no[2:]
        vals["流水號"] = code
        vals["商品料號"] = product_no
        rows.append(vals)
        opt = norm(vals.get("商品選項一"))
        if opt and opt not in seen_options[code]:
            options[code].append(opt)
            seen_options[code].add(opt)
    return rows, options


def build_big_map(big_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(big_path, data_only=True)
    ws = wb["商品資料(大檔)"]
    out: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 5).value)  # E 流水號（目前 reset schema）
        if not code:
            continue
        # 若重複，保留第一個有完整房名/價格的列；人工 91 回推通常只需要流水號原始主列。
        if code in out:
            continue
        out[code] = {
            "row": r,
            "mark": ws.cell(r, 1).value,
            "usage": ws.cell(r, 2).value,
            "code": code,
            "product_no": norm(ws.cell(r, 6).value),
            "cost": ws.cell(r, 11).value,
            "origin_price": ws.cell(r, 19).value,
            "new_price": ws.cell(r, 20).value,
            "full_room_name": norm(ws.cell(r, 26).value),
        }
    return out


def build_fit_map(size_path: Path) -> dict[str, str]:
    wb = load_workbook(size_path, data_only=True)
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        headers = [norm(ws.cell(1, c).value).replace(" ", "").replace("\n", "") for c in range(1, ws.max_column + 1)]
        fit_cols = [i + 1 for i, h in enumerate(headers) if h == "適合尺寸"]
        if not fit_cols:
            continue
        fit_col = fit_cols[0]
        current = ""
        for r in range(2, ws.max_row + 1):
            raw_code = norm(ws.cell(r, 1).value)
            if raw_code:
                current = raw_code
            if not current or current in out:
                continue
            fit = norm(ws.cell(r, fit_col).value)
            if fit:
                out[current] = fit
    return out


def make_row(big: dict[str, Any], option: str, date_code: str, year: int, fit: str) -> list[Any]:
    full_name = big["full_room_name"]
    product_no = big["product_no"]
    if not product_no:
        m = re.search(r"【([^】]+)】", full_name)
        product_no = m.group(1).strip() if m else ""
    return [
        "服裝、內睡衣 >女裝 >上衣 >T恤/帽T",
        "☰ 上衣>T恤、造型上衣",
        full_name,
        999,
        big["origin_price"],
        big["new_price"],
        big["cost"],
        50,
        sales_start_from_mmdd(date_code, year),
        "2041/8/24",
        "一般",
        "",
        "",
        "6592;101268;101269;99006;99007",
        "信用卡一次付款,全家取貨付款,7-11取貨付款,ATM付款,LINE Pay,街口支付,Apple Pay,AFTEE先享後付,悠遊付",
        "有",
        option,
        "",
        product_no,
        "",
        "",
        "", "", "", "", "", "", "", "", "", "",
        fit,
        FEATURE_TEXT,
        "",
        "PUFII-APP",
        "", "", "",
        "常溫",
        "", "",
        "可退貨",
        "關閉",
        "", "", "", "", "",
        0,
        "是",
    ]


def write_workbook(path: Path, rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品資料"
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(norm(ws.cell(1, c).value)) + 2, 8), 35)
    ws.column_dimensions["AG"].width = 60
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 33).alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def compare_rows(manual_rows: list[dict[str, Any]], generated_rows: list[list[Any]]) -> list[dict[str, Any]]:
    gen_dicts = [{h: row[i] for i, h in enumerate(HEADERS)} for row in generated_rows]
    manual_keyed: dict[tuple[str, str], dict[str, Any]] = {}
    for row in manual_rows:
        key = (norm(row.get("商品料號")), norm(row.get("商品選項一")))
        if key not in manual_keyed:
            manual_keyed[key] = row
    gen_keyed = {(norm(row.get("商品料號")), norm(row.get("商品選項一"))): row for row in gen_dicts}
    diffs: list[dict[str, Any]] = []
    for key, mrow in manual_keyed.items():
        grow = gen_keyed.get(key)
        if not grow:
            diffs.append({"key": key, "field": "ROW", "manual": "存在", "generated": "缺列"})
            continue
        for col in COMPARE_COLUMNS:
            mv = norm(mrow.get(col))
            gv = norm(grow.get(col))
            if mv != gv:
                diffs.append({"key": key, "field": col, "manual": mv, "generated": gv})
    for key in gen_keyed:
        if key not in manual_keyed:
            diffs.append({"key": key, "field": "ROW", "manual": "缺列", "generated": "存在"})
    return diffs


def generate(base: Path, manual_91: Path, date_code: str, year: int) -> tuple[Path, Path]:
    big_path = base / "輸入檔" / "商品資料大檔.xlsx"
    size_path = base / "輸入檔" / "尺寸表&試穿報告.xlsx"
    output_path = base / "輸出檔" / f"91-{date_code}新品上架的檔案-人工91回推測試.xlsx"
    report_path = base / "輸出檔" / f"91-{date_code}人工91回推差異報告.json"

    manual_rows, manual_options = parse_manual_91(manual_91)
    codes = list(manual_options.keys())
    big_map = build_big_map(big_path)
    fit_map = build_fit_map(size_path)

    rows: list[list[Any]] = []
    audit_missing: list[dict[str, Any]] = []
    for code in codes:
        big = big_map.get(code)
        if not big:
            audit_missing.append({"code": code, "field": "商品資料(大檔)", "reason": "人工91有此商品，但大檔找不到流水號"})
            continue
        fit = fit_map.get(code, "")
        missing_fields = []
        for field, label in [("full_room_name", "商品名稱"), ("origin_price", "建議售價"), ("new_price", "售價"), ("cost", "成本")]:
            if norm(big.get(field)) == "":
                missing_fields.append(label)
        if not fit:
            missing_fields.append("銷售重點/適合尺寸")
        if missing_fields:
            audit_missing.append({
                "code": code,
                "big_row": big.get("row"),
                "field": "、".join(missing_fields),
                "reason": "人工91回推測試：大檔/尺寸表缺必填欄位",
            })
            continue
        for option in manual_options[code]:
            rows.append(make_row(big, option, date_code, year, fit))
    write_workbook(output_path, rows)
    diffs = compare_rows(manual_rows, rows)
    report = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "manual_91": str(manual_91),
        "output": str(output_path),
        "date_code": date_code,
        "year": year,
        "manual_product_count": len(codes),
        "manual_row_count": len(manual_rows),
        "generated_row_count": len(rows),
        "missing": audit_missing,
        "diff_count": len(diffs),
        "diffs": diffs[:500],
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path, report_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default=str(DEFAULT_BASE))
    parser.add_argument("--manual-91", required=True)
    parser.add_argument("--date", default="0414")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()
    out, report = generate(Path(args.base), Path(args.manual_91), args.date, args.year)
    print(f"已產出：{out}")
    print(f"差異報告：{report}")


if __name__ == "__main__":
    main()
