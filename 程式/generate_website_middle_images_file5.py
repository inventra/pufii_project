#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""從商品資料(大檔).xlsx 的 商品資料(大檔)+大師 產生 {日期}新品-檔案5-中圖合圖.xlsx。"""
from __future__ import annotations

import argparse, datetime as dt, hashlib, json
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path('/Users/openclaw/創旭/帕妃')
SHEET_NAME = '檔案5.購物網站顏色-多小圖匯入-匯入格式'
HEADERS = ['物品編號', '顏色', '物品明細小圖路徑', '物品明細中圖路徑', '物品明細大圖路徑', '購物上架', '排序', '款式顏色', '中途路徑-前半']


def norm(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def to_int(v: Any) -> int | None:
    s = norm(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def today_mmdd() -> str:
    return dt.datetime.now().strftime('%m%d')


def current_year() -> str:
    return dt.datetime.now().strftime('%Y')


def find_big_file(input_dir: Path) -> Path:
    candidates = [
        input_dir / '商品資料(大檔).xlsx',
        input_dir / '商品資料大檔.xlsx',
    ] + sorted(input_dir.glob('*商品資料*大檔*.xlsx'))
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f'找不到商品資料大檔，請放在 {input_dir}')


def load_listing_codes(wb, date_code: str) -> list[str]:
    ws = wb['商品資料(大檔)']
    codes: list[str] = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)
        purpose = norm(ws.cell(r, 2).value)
        code = norm(ws.cell(r, 5).value)
        if mark == date_code and purpose == '上架' and code and code not in seen:
            codes.append(code)
            seen.add(code)
    return codes


def load_rows(wb, codes: list[str], date_code: str, year: str) -> tuple[list[list[Any]], list[dict[str, Any]]]:
    """大師欄位：A 物品編號、J 商店顏色、AC 圖片路徑、AD 色塊號碼、AE 中圖張數。

    檔案5 以「同一流水號 + 商店顏色」去重後，依中圖張數展開 _01.._NN。
    """
    ws = wb['大師']
    code_set = set(codes)
    order = {c: i for i, c in enumerate(codes)}
    records: dict[tuple[str, str], tuple[int, str, str, int, int]] = {}
    missing: list[dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)
        if code not in code_set:
            continue
        color = norm(ws.cell(r, 10).value)  # J 商店顏色
        image_name = norm(ws.cell(r, 29).value)  # AC 圖片路徑
        color_no = to_int(ws.cell(r, 30).value)  # AD 色塊號碼
        image_count = to_int(ws.cell(r, 31).value)  # AE 中圖張數
        if not color:
            missing.append({'row': r, '物品編號': code, 'missing': '商店顏色'})
            continue
        key = (code, color)
        if key in records:
            continue
        miss = []
        if not image_name:
            miss.append('圖片路徑')
        if color_no is None:
            miss.append('色塊號碼')
        if image_count is None:
            miss.append('中圖張數')
        if miss:
            missing.append({
                'row': r,
                '物品編號': code,
                '顏色': color,
                '圖片路徑': image_name,
                '色塊號碼': norm(ws.cell(r, 30).value),
                '中圖張數': norm(ws.cell(r, 31).value),
                'missing': ','.join(miss),
            })
            continue
        records[key] = (r, code, color, image_name, color_no, image_count)

    rows: list[list[Any]] = []
    ordered = sorted(records.items(), key=lambda item: (order[item[0][0]], item[1][0]))
    for _key, (_r, code, color, image_name, color_no, image_count) in ordered:
        prefix = f'https://photo.pufii.com.tw/{year}/{date_code}/{image_name}-{color_no}'
        for idx in range(1, image_count + 1):
            rows.append([
                code,
                color,
                '',
                f'{prefix}_{idx:02d}.jpg',
                '',
                1,
                idx,
                f'{code}-{color}',
                prefix,
            ])
    return rows, missing


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = len(norm(ws.cell(1, c).value))
        for r in range(2, min(ws.max_row, 100) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 70)


def generate(base: Path, date_code: str, source: Path | None = None, year: str | None = None, dry_run: bool = False) -> Path:
    input_dir = base / '輸入檔'
    output_dir = base / '輸出檔' / f'{date_code}官網上架匯入收音機的檔案'
    output_dir.mkdir(parents=True, exist_ok=True)
    source_path = source or find_big_file(input_dir)
    output_path = output_dir / f'{date_code}新品-檔案5-中圖合圖.xlsx'
    year = year or current_year()

    wb_src = load_workbook(source_path, data_only=True)
    listing_codes = load_listing_codes(wb_src, date_code)
    rows, missing = load_rows(wb_src, listing_codes, date_code, year)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    fill = PatternFill('solid', fgColor='D9EAF7')
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    autosize(ws)

    audit = {
        'generated_at': dt.datetime.now().isoformat(timespec='seconds'),
        'script': str(Path(__file__).resolve()),
        'base': str(base),
        'date_code': date_code,
        'year': year,
        'sheet': SHEET_NAME,
        'columns': HEADERS,
        'filter': '商品資料(大檔)!A 記號 = date_code 且 B 用途 = 上架；取 E 流水號',
        'source_sheet': '大師',
        'source_mapping': {
            '物品編號': '大師!A 物品編號',
            '顏色': '大師!J 商店顏色（同流水號+商店顏色去重）',
            '物品明細小圖路徑': '固定空白',
            '物品明細中圖路徑': 'https://photo.pufii.com.tw/{year}/{date}/{大師!AC 圖片路徑}-{大師!AD 色塊號碼}_{序號兩碼}.jpg',
            '物品明細大圖路徑': '固定空白',
            '購物上架': '固定 1',
            '排序': '中圖序號 1..大師!AE 中圖張數',
            '款式顏色': '物品編號-顏色',
            '中途路徑-前半': 'https://photo.pufii.com.tw/{year}/{date}/{大師!AC 圖片路徑}-{大師!AD 色塊號碼}',
        },
        'listing_code_count': len(listing_codes),
        'listing_codes': listing_codes,
        'row_count': len(rows),
        'missing_count': len(missing),
        'missing': missing[:200],
        'inputs': {'商品資料大檔': {'path': str(source_path), 'sha256': sha256(source_path)}},
        'output': str(output_path),
        'dry_run': dry_run,
    }
    if dry_run:
        print(json.dumps(audit, ensure_ascii=False, indent=2))
        return output_path
    wb.save(output_path)
    audit['output_sha256'] = sha256(output_path)
    Path(str(output_path) + '.audit.json').write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    return output_path


def main() -> None:
    ap = argparse.ArgumentParser(description='產生 {日期}新品-檔案5-中圖合圖')
    ap.add_argument('--base', default=str(DEFAULT_BASE))
    ap.add_argument('--date', default=today_mmdd())
    ap.add_argument('--source', default='')
    ap.add_argument('--year', default='')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    out = generate(Path(args.base), args.date, Path(args.source) if args.source else None, args.year or None, args.dry_run)
    if not args.dry_run:
        print(f'已產出：{out}')
        print(f'產出紀錄：{out}.audit.json')


if __name__ == '__main__':
    main()
