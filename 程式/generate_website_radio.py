#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
帕妃 {日期}官網上架匯入收音機的檔案 產生程式

輸入：
- {base}/輸入檔/商品資料大檔.xlsx

輸出：
- {base}/輸出檔/{日期}官網上架匯入收音機的檔案/{日期}新品-檔案2-{名稱}+{大小}.xlsx
- 同路徑 .audit.json

目前商品篩選暫沿用 91 新品上架檔：商品資料(大檔) B 欄「用途」= 上架。
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_BASE = Path("/Users/openclaw/創旭/帕妃")

HEADERS = [
    "第一碼(特價碼)",
    "物品編號",
    "物品名稱",
    "簡介",
    "小圖片路徑(180x180)",
    "原價",
    "價格",
    "物品排序",
    "重量(g)",
    "ps",
    "銷售",
    "尺寸表連結",
    "試穿報告連結",
    "手機版原始碼",
    "原價標",
]


def norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r"\s+", " ", str(v).strip())


def numeric(v: Any) -> float | int | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else float(v)
    s = str(v).strip()
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except Exception:
        return None


def safe_filename_part(v: Any, fallback: str = "未命名") -> str:
    s = norm(v) or fallback
    s = re.sub(r"[\\/:*?\"<>|]+", "-", s)
    return s.strip(" .") or fallback


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def today_mmdd() -> str:
    return dt.datetime.now().strftime("%m%d")


def load_listing_codes(big_path: Path, date_code: str) -> list[str]:
    """商品資料(大檔)：A 欄記號=日期、B 欄用途=上架。"""
    wb = load_workbook(big_path, data_only=True)
    ws = wb["商品資料(大檔)"]
    codes: list[str] = []
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)  # A 記號
        purpose = norm(ws.cell(r, 2).value)  # B 用途（刪除直播日期欄後）
        if mark == date_code and purpose == "上架":
            code = norm(ws.cell(r, 5).value)  # E 流水號（刪除直播日期欄後）
            if code:
                codes.append(code)
    return codes


def build_serial_map(big_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(big_path, data_only=True)
    ws = wb["流水號"]
    data: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(r, 1).value)  # A 流水號
        if not code:
            continue
        weight = numeric(ws.cell(r, 5).value)  # E 重量
        data[code] = {
            "row": r,
            "code": code,
            "name": norm(ws.cell(r, 4).value),  # D 名稱
            "weight": weight,
        }
    return data


def build_price_map(big_path: Path, date_code: str) -> dict[str, dict[str, Any]]:
    """商品資料(大檔)：A 記號=日期、B 用途=上架；S 原價、T 新品價（刪除直播日期欄後）。"""
    wb = load_workbook(big_path, data_only=True)
    ws = wb["商品資料(大檔)"]
    data: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        mark = norm(ws.cell(r, 1).value)  # A 記號
        purpose = norm(ws.cell(r, 2).value)  # B 用途
        code = norm(ws.cell(r, 5).value)  # E 流水號
        if mark == date_code and purpose == "上架" and code:
            data[code] = {
                "row": r,
                "origin_price": ws.cell(r, 19).value,  # S 原價
                "price": ws.cell(r, 20).value,  # T 新品價
            }
    return data


def autosize(ws) -> None:
    for c in range(1, ws.max_column + 1):
        max_len = len(norm(ws.cell(1, c).value))
        for r in range(2, min(ws.max_row, 80) + 1):
            max_len = max(max_len, len(norm(ws.cell(r, c).value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 8), 35)


def generate(base: Path, date_code: str, dry_run: bool = False) -> Path:
    input_dir = base / "輸入檔"
    output_dir = base / "輸出檔" / f"{date_code}官網上架匯入收音機的檔案"
    output_dir.mkdir(parents=True, exist_ok=True)

    big_path = input_dir / "商品資料大檔.xlsx"

    codes = load_listing_codes(big_path, date_code)
    serial_map = build_serial_map(big_path)
    price_map = build_price_map(big_path, date_code)

    output_path = output_dir / f"{date_code}新品-檔案2-名稱+大小.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "商品資料"
    ws.append(HEADERS)

    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    missing: list[dict[str, Any]] = []
    for code in codes:
        serial = serial_map.get(code)
        price = price_map.get(code)
        if not serial:
            missing.append({"code": code, "field": "流水號分頁", "reason": "找不到對應流水號"})
        if not price:
            missing.append({"code": code, "field": "商品資料(大檔)價格", "reason": "找不到對應流水號/日期/用途"})

        weight = serial.get("weight") if serial else None
        weight_plus_30 = "" if weight is None else weight + 30

        ws.append([
            "Z",  # 第一碼(特價碼)
            serial.get("code", code) if serial else code,  # 物品編號
            serial.get("name", "") if serial else "",  # 物品名稱
            "",  # 簡介
            "",  # 小圖片路徑(180x180)
            price.get("origin_price", "") if price else "",  # 原價：商品資料(大檔) S 欄
            price.get("price", "") if price else "",  # 價格：商品資料(大檔) T 欄
            0,  # 物品排序
            weight_plus_30,  # 重量(g)
            "",  # ps
            1,  # 銷售
            "",  # 尺寸表連結
            "",  # 試穿報告連結
            "",  # 手機版原始碼
            1,  # 原價標
        ])

    ws.freeze_panes = "A2"
    autosize(ws)

    audit = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "script": str(Path(__file__).resolve()),
        "base": str(base),
        "date_code": date_code,
        "folder": str(output_dir),
        "sheet": "商品資料",
        "filter": "商品資料(大檔)!A 記號 = date_code 且 B 用途 = 上架",
        "column_count": len(HEADERS),
        "last_column": get_column_letter(len(HEADERS)),
        "last_header": HEADERS[-1],
        "row_count": len(codes),
        "inputs": {"商品資料大檔.xlsx": {"path": str(big_path), "sha256": sha256(big_path)}},
        "output": str(output_path),
        "missing_or_pending": missing,
        "dry_run": dry_run,
    }

    if dry_run:
        print(json.dumps(audit, ensure_ascii=False, indent=2))
        return output_path

    wb.save(output_path)
    audit["output_sha256"] = sha256(output_path)
    Path(str(output_path) + ".audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="產生 {日期}官網上架匯入收音機的檔案")
    parser.add_argument("--base", default=str(DEFAULT_BASE), help="帕妃資料夾根目錄")
    parser.add_argument("--date", default=today_mmdd(), help="日期 MMDD，例如 0414；未填預設今天")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    output = generate(Path(args.base), args.date, args.dry_run)
    if not args.dry_run:
        print(f"已產出：{output}")
        print(f"產出紀錄：{output}.audit.json")


if __name__ == "__main__":
    main()
